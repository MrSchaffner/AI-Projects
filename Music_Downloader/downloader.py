import subprocess
import sys
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SPREADSHEET = "songs.xlsx"
DOWNLOAD_DIR = "downloads"
HEADERS = ["Artist", "Song", "Status", "Notes"]
STATUS_WAITING = "waiting"
STATUS_DOWNLOADED = "downloaded"
STATUS_NOT_FOUND = "not found"
STATUS_ERROR = "error"

STATUS_COLORS = {
    STATUS_DOWNLOADED: "C6EFCE",
    STATUS_NOT_FOUND:  "FFEB9C",
    STATUS_ERROR:      "FFC7CE",
    STATUS_WAITING:    "FFFFFF",
}


def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Songs"

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40

    sample_songs = [
        ("Radiohead", "Creep", STATUS_WAITING, ""),
        ("Nirvana", "Smells Like Teen Spirit", STATUS_WAITING, ""),
    ]
    for row_data in sample_songs:
        ws.append(row_data)

    wb.save(SPREADSHEET)
    print(f"Created {SPREADSHEET} with sample songs. Edit it and run again.")


def load_songs():
    wb = load_workbook(SPREADSHEET)
    ws = wb.active
    songs = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        artist = row[0].value
        song   = row[1].value
        status = row[2].value if row[2].value else STATUS_WAITING
        if artist and song and status != STATUS_DOWNLOADED:
            songs.append((row[0].row, artist.strip(), song.strip(), status))
    return wb, ws, songs


def update_row(ws, row_num, status, notes=""):
    ws.cell(row=row_num, column=3, value=status)
    ws.cell(row=row_num, column=4, value=notes)

    color = STATUS_COLORS.get(status, "FFFFFF")
    fill  = PatternFill("solid", start_color=color)
    for col in range(1, 5):
        ws.cell(row=row_num, column=col).fill = fill


def download_song(artist, song):
    query = f"ytsearch1:{artist} - {song}"
    Path(DOWNLOAD_DIR).mkdir(exist_ok=True)
    output_template = str(Path(DOWNLOAD_DIR) / "%(title)s.%(ext)s")

    try:
        result = subprocess.run(
            [
                sys.executable, "-m", "yt_dlp",
                query,
                "--extract-audio",
                "--audio-format", "mp3",
                "--audio-quality", "0",
                "--output", output_template,
                "--no-playlist",
                "--quiet",
                "--no-warnings",
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        return STATUS_ERROR, "Timed out after 120 seconds"

    if result.returncode == 0:
        return STATUS_DOWNLOADED, ""

    stderr = result.stderr.strip()
    last_line = stderr.splitlines()[-1] if stderr else "Unknown error"

    if "No video formats" in last_line or "Unable to extract" in last_line:
        return STATUS_NOT_FOUND, last_line[:100]
    return STATUS_ERROR, last_line[:100]


def main():
    if not Path(SPREADSHEET).exists():
        create_template()
        return

    wb, ws, songs = load_songs()

    if not songs:
        print("No songs left to download. All are marked 'downloaded'.")
        return

    print(f"Found {len(songs)} song(s) to process.\n")

    for row_num, artist, song, status in songs:
        label = f"{artist} - {song}"
        print(f"  Downloading: {label} ... ", end="", flush=True)
        new_status, notes = download_song(artist, song)
        update_row(ws, row_num, new_status, notes)
        print(new_status)

    wb.save(SPREADSHEET)
    print(f"\nDone. Spreadsheet updated: {SPREADSHEET}")
    print(f"MP3s saved to: {DOWNLOAD_DIR}/")


if __name__ == "__main__":
    main()